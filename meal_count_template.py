import calendar
import datetime
from openpyxl import load_workbook
from excel_utils import get_excel_column_label, write_to_cell_excel
from useful_utils import is_valid_year
from color_utils import (
    odd_header_color,
    odd_header_fill,
    even_header_color,
    even_header_fill,
)
from config import (
    failure_string,
    success_string,
    info_string,
    workbook_name,
    meal_count_sheet_name,
)

month_names = {
    1: "January",
    2: "February",
    3: "March",
    4: "April",
    5: "May",
    6: "June",
    7: "July",
    8: "August",
    9: "September",
    10: "October",
    11: "November",
    12: "December",
}

sheet_name = meal_count_sheet_name
boarders_list_names = "boarders.txt"


def init_meal_sheet(workbook_openpyxl=None):
    ## get name of the sheet
    sheet = sheet_name
    print(f"{info_string}Generating Meal Sheet Template")
    change_sheet_name = input(
        f"What should be the sheet name? [default: {sheet_name}] blank for default | "
    )
    if change_sheet_name:
        sheet = change_sheet_name
    print(
        f"{info_string}IMPORTANT: if sheet named '{sheet}' is present in {workbook_name} file, it will be replaced."
    )
    ## take year input
    current_year = datetime.date.today().year
    change_year = input(
        f"For which YEAR, You want to create '{sheet}' sheet? [default:{current_year}] blank for default | "
    )
    year = current_year
    if change_year:
        if is_valid_year(change_year):
            year = int(change_year)
        else:
            raise Exception("Not a Valid Year.")
    ## take month input
    current_month_no = datetime.date.today().month
    print()
    print(f"For which month(of {year}), You want to create {sheet_name} sheet: ")
    for no, month in month_names.items():
        print(no, ":", month)
    print()
    month_no = input("Enter a number[Blank for current month]: ")
    if not month_no:
        month_no = current_month_no
    else:
        month_no = int(month_no)

    month = month_no
    days_in_month = calendar.monthrange(year, month)[1]
    ## write to excel
    workbook = None
    if workbook_openpyxl is not None:
        workbook = workbook_openpyxl
    else:
        workbook = load_workbook(workbook_name)
    worksheet_name = sheet
    # Check if the worksheet already exists
    if worksheet_name not in workbook.sheetnames:
        # If it doesn't exist, create a new worksheet
        worksheet = workbook.create_sheet(title=worksheet_name)
    else:
        # If it exists, delete it then create one
        worksheet = workbook[worksheet_name]
        workbook.remove(worksheet)
        worksheet = workbook.create_sheet(title=worksheet_name)

    # writing
    sheet = worksheet
    sheet.merge_cells("A1:E1")
    write_to_cell_excel(
        sheet,
        "A1",
        f"({month_names[month]}, {year}) date->",
        bold=True,
        center_aligned=True,
    )
    write_to_cell_excel(
        sheet,
        "A2",
        "Name",
        bold=True,
        center_aligned=True,
        color="#ffffff",
        fill="#000000",
    )
    # status
    write_to_cell_excel(
        sheet,
        "B2",
        "Status",
        bold=True,
        center_aligned=True,
        color="#000000",
        fill="#ffffff",
    )
    # rules
    write_to_cell_excel(
        sheet,
        "C2",
        "Rules",
        bold=True,
        center_aligned=True,
        color="#ffffff",
        fill="#000000",
    )
    # deposit
    write_to_cell_excel(
        sheet,
        "D2",
        "Deposit",
        bold=True,
        center_aligned=True,
        color="#ffffff",
        fill="#000000",
    )
    # accessories charge
    write_to_cell_excel(
        sheet,
        "E2",
        "Acs Charge",
        bold=True,
        center_aligned=True,
        color="#ffffff",
        fill="#000000",
    )

    prev_col = 1
    for d in range(days_in_month):
        cell_label = d + 4 if d + 1 == 1 else prev_col + 6
        prev_col = cell_label
        column_label_day = get_excel_column_label(cell_label + 2)
        column_label_day_egg = get_excel_column_label(cell_label + 3)
        column_label_day_guest = get_excel_column_label(cell_label + 4)
        column_label_night = get_excel_column_label(cell_label + 5)
        column_label_night_egg = get_excel_column_label(cell_label + 6)
        column_label_night_guest = get_excel_column_label(cell_label + 7)
        # colors
        header_fill = odd_header_fill if d % 2 == 0 else even_header_fill
        header_color = odd_header_color if d % 2 == 0 else even_header_color

        sheet.merge_cells(f"{column_label_day}1:{column_label_night_guest}1")
        write_to_cell_excel(
            sheet,
            f"{column_label_day}1",
            d + 1,
            True,
            True,
            fill=header_fill,
            color=header_color,
        )
        write_to_cell_excel(
            sheet,
            f"{column_label_day}2",
            "day",
            center_aligned=True,
            fill=header_fill,
            color=header_color,
        )
        write_to_cell_excel(
            sheet,
            f"{column_label_day_egg}2",
            "egg(day)",
            center_aligned=True,
            fill=header_fill,
            color=header_color,
        )
        write_to_cell_excel(
            sheet,
            f"{column_label_day_guest}2",
            "guest(day)",
            center_aligned=True,
            fill=header_fill,
            color=header_color,
        )
        write_to_cell_excel(
            sheet,
            f"{column_label_night}2",
            "night",
            center_aligned=True,
            fill=header_fill,
            color=header_color,
        )
        write_to_cell_excel(
            sheet,
            f"{column_label_night_egg}2",
            "egg(night)",
            center_aligned=True,
            fill=header_fill,
            color=header_color,
        )
        write_to_cell_excel(
            sheet,
            f"{column_label_night_guest}2",
            "guest(night)",
            center_aligned=True,
            fill=header_fill,
            color=header_color,
        )
    write_names = input(
        f"Write names from {boarders_list_names} file? (Y/y or blank for no) "
    )
    if write_names == "Y" or write_names == "y":
        # reading names
        with open(boarders_list_names) as f:
            contents = f.read()
        boarders = contents.splitlines()
        # write names
        print()
        print("Writing Names-")
        for i, name in enumerate(boarders, start=3):
            write_to_cell_excel(sheet, f"A{i}", name, center_aligned=True)
            print(f"   {name}")
        print("Names written")
        print()

    # set column size
    sheet.column_dimensions["A"].width = 23
    sheet.column_dimensions["C"].width = 15
    sheet.column_dimensions["D"].width = 15
    sheet.column_dimensions["E"].width = 12

    # freeze panes
    sheet.freeze_panes = sheet["B3"]

    workbook.save(workbook_name)
    print(f"{success_string}Template generated.")
    workbook.close()


# init_meal_sheet()
