import calendar
from openpyxl import load_workbook
from excel_utils import write_to_cell_excel

import config

# from config import (
#     failure_string,
#     success_string,
#     info_string,
#     workbook_name,
#     meal_count_sheet_name,
# )

month_names = {
    1: "January",
    2: "February",
    3: "March",
    4: "April",
    5: "May",
    6: "June",
    7: "July",
    8: "August",
    9: "September",
    10: "October",
    11: "November",
    12: "December",
}

sheet_name = config.establishment_sheet_name


def init_establishment_sheet(workbook_openpyxl=None):
    ## get name of the sheet
    sheet = sheet_name
    print(f"{config.info_string}Generating Establishment Sheet Template.")
    print(
        f"{config.info_string}IMPORTANT: if sheet named '{sheet}' is present in {config.workbook_name} file, it will be replaced."
    )
    ## write to excel
    workbook = None
    if workbook_openpyxl is not None:
        workbook = workbook_openpyxl
    else:
        workbook = load_workbook(config.workbook_name)
    worksheet_name = sheet
    # Check if the worksheet already exists
    if worksheet_name not in workbook.sheetnames:
        # If it doesn't exist, create a new worksheet
        worksheet = workbook.create_sheet(title=worksheet_name)
    else:
        # If it exists, delete it then create one
        worksheet = workbook[worksheet_name]
        workbook.remove(worksheet)
        worksheet = workbook.create_sheet(title=worksheet_name)

    # writing
    sheet = worksheet
    ######## Daily Marketing
    sheet.merge_cells("A1:B1")  # daily marketing
    write_to_cell_excel(
        sheet,
        "A1",
        "Establishment Money",
        bold=True,
        center_aligned=True,
        color="#ffffff",
        fill="#000000",
    )
    write_to_cell_excel(
        sheet,
        "A2",
        "Title",
        bold=True,
        center_aligned=True,
        color="#023e8a",
        fill_reset=True,
    )
    write_to_cell_excel(
        sheet,
        "B2",
        "Amount",
        bold=True,
        center_aligned=True,
        color="#023e8a",
        fill_reset=True,
    )
    # set column size
    sheet.column_dimensions["A"].width = 40
    sheet.column_dimensions["B"].width = 15

    # freeze panes
    sheet.freeze_panes = sheet["A3"]

    workbook.save(config.workbook_name)
    print(f"{config.success_string}Marketing Template generated.")
    workbook.close()
