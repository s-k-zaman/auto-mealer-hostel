from openpyxl import load_workbook
from excel_utils import get_excel_column_label, write_to_cell_excel
import config


month_names = {
    1: "January",
    2: "February",
    3: "March",
    4: "April",
    5: "May",
    6: "June",
    7: "July",
    8: "August",
    9: "September",
    10: "October",
    11: "November",
    12: "December",
}

sheet_name = config.meal_routine_sheet_name


def init_routine_meal_sheet(month, year, routine_items_list, workbook_openpyxl=None):
    ## get name of the sheet
    sheet = sheet_name
    print(
        f"{config.info_string}Generating Routine Meal Sheet Template for month of {month}, {year}."
    )
    print(
        f"{config.info_string}IMPORTANT: if sheet named '{sheet}' is present in {config.workbook_name} file, it will be replaced."
    )

    ## write to excel
    workbook = None
    if workbook_openpyxl is not None:
        workbook = workbook_openpyxl
    else:
        workbook = load_workbook(config.workbook_name)
    worksheet_name = sheet
    # Check if the worksheet already exists
    if worksheet_name not in workbook.sheetnames:
        # If it doesn't exist, create a new worksheet
        worksheet = workbook.create_sheet(title=worksheet_name)
    else:
        # If it exists, delete it then create one
        worksheet = workbook[worksheet_name]
        workbook.remove(worksheet)
        worksheet = workbook.create_sheet(title=worksheet_name)

    # writing
    sheet = worksheet

    # putting all the dates
    starting_cell = 1
    for item_name in routine_items_list:
        # colors
        fill_color = "#ffffff"
        text_color = "#000000"
        if item_name == "beef":
            fill_color = "#A52A2A"
            text_color = "#ffffff"
        if item_name == "egg":
            fill_color = "#FFD700"
            text_color = "#000000"
        if item_name == "veg":
            fill_color = "#008000"
            text_color = "#ffffff"
        if item_name == "fish":
            fill_color = "#0000FF"
            text_color = "#ffffff"
        if item_name == "chicken":
            fill_color = "#FF4500"
            text_color = "#000000"

        cell_col = get_excel_column_label(starting_cell)
        # write name
        cell = f"{cell_col}1"
        write_to_cell_excel(
            sheet,
            cell,
            item_name,
            center_aligned=True,
            fill=fill_color,
            color=text_color,
        )
        sheet.column_dimensions[cell_col].width = 25
        starting_cell += 1
        # write date
        cell_col = get_excel_column_label(starting_cell)
        cell = f"{cell_col}1"
        write_to_cell_excel(
            sheet,
            cell,
            "dates",
            center_aligned=True,
            fill=fill_color,
            color=text_color,
        )
        sheet.column_dimensions[cell_col].width = 15
        starting_cell += 1

    # freeze panes
    sheet.freeze_panes = sheet["A2"]

    workbook.save(config.workbook_name)
    print(f"{config.success_string}Meal Routine Template generated.")
    workbook.close()
