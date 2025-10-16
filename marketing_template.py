import calendar
from openpyxl import load_workbook
from excel_utils import write_to_cell_excel

import config

# from config import (
#     failure_string,
#     success_string,
#     info_string,
#     workbook_name,
#     meal_count_sheet_name,
# )

month_names = {
    1: "January",
    2: "February",
    3: "March",
    4: "April",
    5: "May",
    6: "June",
    7: "July",
    8: "August",
    9: "September",
    10: "October",
    11: "November",
    12: "December",
}

sheet_name = config.marketing_sheet_name


def init_marketing_sheet(month, year, workbook_openpyxl=None):
    ## get name of the sheet
    sheet = sheet_name
    print(
        f"{config.info_string}Generating Marketing Sheet Template for month of {month}, {year}."
    )
    print(
        f"{config.info_string}IMPORTANT: if sheet named '{sheet}' is present in {config.workbook_name} file, it will be replaced."
    )

    days_in_month = calendar.monthrange(year, month)[1]

    ## write to excel
    workbook = None
    if workbook_openpyxl is not None:
        workbook = workbook_openpyxl
    else:
        workbook = load_workbook(config.workbook_name)
    worksheet_name = sheet
    # Check if the worksheet already exists
    if worksheet_name not in workbook.sheetnames:
        # If it doesn't exist, create a new worksheet
        worksheet = workbook.create_sheet(title=worksheet_name)
    else:
        # If it exists, delete it then create one
        worksheet = workbook[worksheet_name]
        workbook.remove(worksheet)
        worksheet = workbook.create_sheet(title=worksheet_name)

    # writing
    sheet = worksheet
    ######## Daily Marketing
    sheet.merge_cells("A1:D1")  # daily marketing
    write_to_cell_excel(
        sheet,
        "A1",
        f"Daily marketing ({month_names[month]}, {year})",
        bold=True,
        center_aligned=True,
        color="#ffffff",
        fill="#000000",
    )
    write_to_cell_excel(
        sheet,
        "A2",
        "Date",
        bold=True,
        center_aligned=True,
        color="#023e8a",
        fill_reset=True,
    )
    write_to_cell_excel(
        sheet,
        "B2",
        "Name",
        bold=True,
        center_aligned=True,
        color="#023e8a",
        fill_reset=True,
    )
    write_to_cell_excel(
        sheet,
        "C2",
        "Money Spent",
        bold=True,
        center_aligned=True,
        color="#023e8a",
        fill_reset=True,
    )
    write_to_cell_excel(
        sheet,
        "D2",
        "Remarks",
        bold=True,
        center_aligned=True,
        color="#023e8a",
        fill_reset=True,
    )

    # putting all the dates
    starting_row = 3
    for d in range(1, days_in_month + 1):
        cell = f"A{starting_row}"
        date_to_write = f"{d}/{month}/{year}"
        write_to_cell_excel(
            sheet,
            cell,
            date_to_write,
            center_aligned=True,
            fill_reset=True,
            color_reset=True,
        )
        starting_row += 1

    # set column size
    sheet.column_dimensions["A"].width = 15
    sheet.column_dimensions["B"].width = 23
    sheet.column_dimensions["C"].width = 15
    sheet.column_dimensions["D"].width = 12

    ######## Extra Marketing
    sheet.merge_cells("G1:I1")
    write_to_cell_excel(
        sheet,
        "G1",
        f"Extra marketing ({month_names[month]}, {year})",
        bold=True,
        center_aligned=True,
        color="#ffffff",
        fill="#000000",
    )
    write_to_cell_excel(
        sheet,
        "G2",
        "Date",
        bold=True,
        center_aligned=True,
        color="#023e8a",
        fill_reset=True,
    )
    write_to_cell_excel(
        sheet,
        "H2",
        "Name",
        bold=True,
        center_aligned=True,
        color="#023e8a",
        fill_reset=True,
    )
    write_to_cell_excel(
        sheet,
        "I2",
        "Money Spent",
        bold=True,
        center_aligned=True,
        color="#023e8a",
        fill_reset=True,
    )
    sheet.column_dimensions["G"].width = 15
    sheet.column_dimensions["H"].width = 23
    sheet.column_dimensions["I"].width = 15

    ######## Extra Rice
    sheet.merge_cells("K1:M1")
    write_to_cell_excel(
        sheet,
        "K1",
        f"Extra Rice + other spent({month_names[month]}, {year})",
        bold=True,
        center_aligned=True,
        color="#ffffff",
        fill="#000000",
    )
    write_to_cell_excel(
        sheet,
        "K2",
        "Date",
        bold=True,
        center_aligned=True,
        color="#023e8a",
        fill_reset=True,
    )
    write_to_cell_excel(
        sheet,
        "L2",
        "Name",
        bold=True,
        center_aligned=True,
        color="#023e8a",
        fill_reset=True,
    )
    write_to_cell_excel(
        sheet,
        "M2",
        "Price",
        bold=True,
        center_aligned=True,
        color="#023e8a",
        fill_reset=True,
    )
    sheet.column_dimensions["K"].width = 15
    sheet.column_dimensions["L"].width = 23
    sheet.column_dimensions["M"].width = 15

    # freeze panes
    sheet.freeze_panes = sheet["A3"]

    workbook.save(config.workbook_name)
    print(f"{config.success_string}Marketing Template generated.")
    workbook.close()


# init_meal_sheet()
