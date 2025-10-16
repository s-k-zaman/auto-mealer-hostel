import pandas as pd
import pendulum
from openpyxl import load_workbook
from pandas.io.formats.format import math

import config
from boarder import Boarder
from color_utils import grand_meal_color, grand_meal_fill
from establishment_template import init_establishment_sheet
from excel_utils import (
    get_cell_for_boarder_from_df,
    get_excel_column_label,
    get_val_from_excel_file,
    is_sheet_present_excel,
    write_in_excel_file,
    write_to_cell_excel,
)
from marketing_template import init_marketing_sheet
from meal_count_template import init_meal_sheet
from routine_meal_template import init_routine_meal_sheet
from useful_utils import (
    check_meal_alternate,
    choose_a_date,
    choose_time,
    final_cofirm,
    fuzzy_select_from_list,
    get_date_color,
    get_date_fill,
    get_date_time_next_meal_counting,
    get_date_time_now,
    get_meal_sheet_label_props,
    get_start_to_end_date_time,
    print_in_2_cols,
    select_a_item_with_default,
    select_items_from_list,
    take_input,
)

if not config.debugging:
    pd.options.mode.chained_assignment = None  # default='warn'


class Hostel:
    def __init__(self, filename) -> None:
        self.filename = filename

        self.selected_boarder_row = None
        self.selected_boarder = None
        self.total_boarders = 0
        self.month = None
        self.month_string = None
        self.month_days = None
        self.invalid_cells = [[]]

        self.parse_file()

    def parse_file(self):
        excel_file = self.filename
        # if meal_count_sheet_name is not present create it.
        sheet_present = is_sheet_present_excel(config.meal_count_sheet_name)
        if not sheet_present:
            print(f"{config.failure_string} Meal count sheet not present.")
            confirmed = final_cofirm("Want to create meal sheet Template?")
            if confirmed:
                init_meal_sheet()
            else:
                raise Exception(
                    f"{config.failure_string}Can not proceed without mealsheet file."
                )

        # parse routine
        self.routine = pd.DataFrame(config.routine[1:], columns=config.routine[0])
        self.all_routine_values = list(
            set(self.routine[["day", "night"]].values.flatten())
        )

        work_book = load_workbook(excel_file)
        work_sheet = work_book[config.meal_count_sheet_name]
        # get header rows
        header_rows = work_sheet.iter_rows(min_row=1, max_row=2, values_only=True)
        rows = []
        for row in header_rows:
            rows.append(row)
        header_with_date = rows[0]
        header_with_label = rows[1]
        # getting month
        self.month_string = header_with_date[0].split(",")[0].split("(")[-1].lower()
        self.month = pendulum.parse(self.month_string, strict=False).month
        # getting Year
        year = header_with_date[0].split(",")[-1].split(")")[0]
        year = int(year)
        self.year = year
        ## get day labels and make headers with it, to work with pandas
        labels_count = 0
        start_counter = False
        for date in header_with_date:
            if date == 2:
                break
            if date == 1:
                start_counter = True
            if start_counter:
                labels_count += 1
        df_label = []
        date = None
        for i in range(len(header_with_label)):
            if isinstance(header_with_date[i], int):
                date = header_with_date[i]
            if date is not None:
                df_label.append(f"{date}_{header_with_label[i]}")
            else:
                df_label.append(header_with_label[i])
        self.boarders_df_labels = df_label
        boarders_df = pd.read_excel(
            io=self.filename,
            sheet_name=config.meal_count_sheet_name,
            names=df_label,
            header=None,
            engine="openpyxl",
        )
        # reset index to start from 1 not 0.
        boarders_df.index = boarders_df.index + 1
        self.boarders_df = boarders_df.iloc[2:]
        # set total boarder count
        self.total_boarders = len(self.boarders_df)
        meal_sheet_label_props = get_meal_sheet_label_props(
            list(self.boarders_df.columns)
        )
        self.month_days = meal_sheet_label_props["days"]
        # get all the meal related datas
        self.total_meals = self.get_total_meal_of_motnth()
        self.total_guest_meals = self.get_total_guest_meals()
        self.total_grand_guest_meals = self.get_total_grand_guest_meals()
        self.total_deposit = self.get_total_deposits()
        self.total_eggs = self.get_total_eggs()
        self.total_acs_charge = self.get_total_accessories_charge()
        # check for any invalid datas
        self.invalid_cells = [[]]
        self.check_invalid_cells()
        if len(self.invalid_cells[:-1]):
            print(
                f"{config.failure_string} INVALID DATA PRESENT IN THE FILE, PLEASE CORRECT IT [select - check for invalid cells, to see cells]."
            )
        self.all_boarders = self.boarders_df["Name"].tolist()
        if self.selected_boarder:
            self.select_boarder(self.selected_boarder.name, show_selected_msg=False)

        self.is_marketing_sheet = is_sheet_present_excel(config.marketing_sheet_name)
        self.is_establishment_sheet = is_sheet_present_excel(
            config.establishment_sheet_name
        )
        if not self.is_marketing_sheet:
            print("Generating Markeing sheet.")
            init_marketing_sheet(self.month, self.year)
        if not self.is_establishment_sheet:
            print("Generating Establishment Sheet")
            init_establishment_sheet()
        self.parse_marketing()

        # if routine meal sheet not present create it
        self.is_meal_routine_sheet = is_sheet_present_excel(
            config.meal_routine_sheet_name
        )
        if not self.is_meal_routine_sheet:
            print("Generate Meal Routine sheet.")
            init_routine_meal_sheet(self.month, self.year, self.all_routine_values)
        self.parse_routine_meal()
        # calculations
        self.calculate_establishment_charge()
        self.calculate_meal_charge()

        work_book.close()

    def parse_routine_meal(self):
        excel_file = self.filename
        self.routine_meal_df = pd.read_excel(
            io=excel_file,
            sheet_name=config.meal_routine_sheet_name,
            engine="openpyxl",
        )

        final_cols = []
        headers = self.routine_meal_df.columns.tolist()
        prev_head = None
        for head in headers:
            if "dates" in head:
                final_cols.append(f"{prev_head}-dates")
                continue
            else:
                prev_head = head
            final_cols.append(head)
        self.routine_meal_df.columns = final_cols

    def parse_marketing(self):
        excel_file = self.filename
        self.marketing_df = pd.read_excel(
            io=excel_file,
            sheet_name=config.marketing_sheet_name,
            engine="openpyxl",
            header=1,
            usecols=range(4),
        )
        is_string = self.marketing_df["Remarks"].map(lambda x: isinstance(x, str)).any()
        if is_string:
            self.marketing_df["Remarks"] = self.marketing_df["Remarks"].str.lower()
            # remove rows where remarks is na
            self.marketing_df = self.marketing_df[
                ~self.marketing_df["Remarks"].isin(["na", "n/a", "not allowed"])
            ]
        self.total_marketing = self.marketing_df["Money Spent"].sum()
        ## extra marketings
        self.extra_marketing_df = pd.read_excel(
            io=excel_file,
            sheet_name=config.marketing_sheet_name,
            engine="openpyxl",
            header=1,
            usecols=range(6, 9),
        )
        self.extra_marketing_df.columns = [
            col.split(".")[0] for col in list(self.extra_marketing_df.columns)
        ]
        self.total_extra_marketing = self.extra_marketing_df["Money Spent"].sum()
        ## extra rice
        self.extra_rice_df = pd.read_excel(
            io=excel_file,
            sheet_name=config.marketing_sheet_name,
            engine="openpyxl",
            header=1,
            usecols=range(10, 13),
        )
        self.extra_rice_df.columns = [
            col.split(".")[0] for col in list(self.extra_rice_df.columns)
        ]
        self.total_extra_rice = self.extra_rice_df["Price"].sum()

    def main_menu(self, quit_str="000"):
        quit_str = str(quit_str)

        def refresh():
            return

        auto_meal_txt = "Do meal count(automatic)"
        options_table = [
            dict(view="Select a Boarder", function=self.work_with_a_boarder),
            dict(view="Set grand meal day.", function=self.set_grand_meal_day),
            dict(view="Remove grand meal day.", function=self.remove_grand_meal_day),
            dict(view="Add a boarder.", function=self.add_boarder),
            dict(view="See meal count details", function=self.print_meal_details),
            dict(
                view="See deposit/meal details", function=self.see_low_budget_boarders
            ),
            dict(view="See routine", function=self.print_routine),
            dict(view="Refresh.", function=refresh),
            dict(view="Check Invalid cells.", function=self.show_invalid_cells),
            dict(view="see final meals", function=self.print_final_meals),
            dict(view="produce final sheet", function=self.write_final_sheet),
            dict(view=auto_meal_txt, function=self.do_next_meal_count),
        ]
        # informations to display
        informations_to_show = self.get_informations_to_show()
        options = [item["view"].strip() for item in options_table]
        choice = None
        date_now, _ = get_date_time_now()
        day_today = pendulum.today().format("dddd")
        while True:
            next_meal = self.allow_next_meal_count()
            if not next_meal and auto_meal_txt in options:
                options.remove(auto_meal_txt)
            elif next_meal and auto_meal_txt not in options:
                options.append(auto_meal_txt)

            choice = select_items_from_list(
                options,
                multi_select=False,
                quit_str=quit_str,
                quit_text="EXIT",
                presentation_text=f"Main Menu ({self.month_string.title()}-{date_now}, {day_today}) | Manager: {config.MANAGER_NAME}",
                presentation_symbol="#",
                width=50,
                informations_to_show=informations_to_show,
            )
            if choice is None:
                break
            choice_idx = options.index(choice)
            options_table[choice_idx]["function"]()
            self.parse_file()
            informations_to_show = self.get_informations_to_show()
            date_now, _ = get_date_time_now()
            day_today = pendulum.today().format("dddd")

    def print_final_meals(self):
        total_due = 0
        total_refund = 0
        for b in self.all_boarders:
            self.select_boarder(b, show_selected_msg=False)
            total_charge = (
                (self.selected_boarder.total_meal * self.meal_charge)
                + (self.selected_boarder.guest_meals * config.guest_meal_price)
                + (self.selected_boarder.total_eggs * config.egg_price)
                + (self.selected_boarder.grand_guest_meals * config.grand_guest_price)
                + self.selected_boarder.accessories_charge
                + self.establishment_charge
            )

            remaining = total_charge - self.selected_boarder.deposit
            if remaining > 0:
                total_due += remaining
            if remaining < 0:
                total_refund += abs(remaining)

            print(
                f"{self.selected_boarder.name}- total charge: {total_charge:.2f} | deposit:{self.selected_boarder.deposit} | remaining:{remaining:.2f}"
            )
        remaining_amount = self.get_total_deposits() - (
            self.total_marketing + self.total_extra_marketing + self.total_extra_rice
        )
        final_collection = (total_due - total_refund) + remaining_amount
        print()
        print(
            f"total due:{total_due:.2f} | total refund:{total_refund:.2f} | remaining amount:{remaining_amount} | final collection:{round(final_collection)}"
        )
        input()

    def print_routine(self):
        print(
            """
################################
#         Meal Routine         #
################################
"""
        )
        print(self.routine)
        input()

    def work_with_a_boarder(self):
        self.select_boarder()
        if self.selected_boarder:
            self.selected_boarder.menu()

    ############# Read Data
    def check_invalid_cells(self):
        boarders_list = self.boarders_df["Name"].values
        # check Number cells
        headers = ["Acs Charge", "Deposit"]
        for label in self.boarders_df_labels:
            to_check_for = ["guest", "egg"]
            present = any([sub_str in label for sub_str in to_check_for])
            if present:
                headers.append(label)
        numbers_df = self.boarders_df[headers]
        numbers_df_values = numbers_df.values
        for i, boarder_values in enumerate(numbers_df_values):
            name = boarders_list[i]
            for j, value in enumerate(boarder_values):
                label = headers[j]
                if not isinstance(value, (int, float)):
                    cell = get_cell_for_boarder_from_df(self.boarders_df, name, label)
                    cell_label = label.replace("_", " - ")
                    self.add_to_invalid_cell(name, cell, "Numbers", cell_label, value)

        # check for string values
        headers = ["Name", "Status", "Rules"]
        for label in self.boarders_df_labels:
            to_check_for = ["day", "night"]
            present = any([label.endswith(sub_str) for sub_str in to_check_for])
            if present:
                headers.append(label)
        strings_df = self.boarders_df[headers]
        string_df_values = strings_df.values
        for i, boarder_values in enumerate(string_df_values):
            name = boarders_list[i]
            for j, value in enumerate(boarder_values):
                label = headers[j]
                # if it is blank
                if not isinstance(value, str):
                    if math.isnan(value):
                        continue
                # if it is a number
                if isinstance(value, (int, float)):
                    cell = get_cell_for_boarder_from_df(self.boarders_df, name, label)
                    cell_label = label.replace("_", " - ")
                    self.add_to_invalid_cell(name, cell, "Texts", cell_label, value)
                # apply individual rules.
                if label == headers[1]:
                    # for status
                    allowed_values = (
                        config.abbr_status_on
                        + config.abbr_status_off
                        + config.abbr_continue_guest_meal
                    )
                    allowed_values_string = ", ".join(allowed_values)
                    allowed_values_string += (
                        ", and one Number(>0) represents guest meal for auto count."
                    )
                    sub_values = value.split(",")
                    sub_values = [val.strip() for val in sub_values]
                    have_digit = False
                    for val in sub_values:
                        if val.isdigit():
                            if int(val) <= 0 and have_digit:
                                cell = get_cell_for_boarder_from_df(
                                    self.boarders_df, name, label
                                )
                                cell_label = label.replace("_", " - ")
                                self.add_to_invalid_cell(
                                    name, cell, allowed_values_string, cell_label, val
                                )
                            have_digit = True
                            continue
                        if val not in allowed_values:
                            cell = get_cell_for_boarder_from_df(
                                self.boarders_df, name, label
                            )
                            cell_label = label.replace("_", " - ")
                            self.add_to_invalid_cell(
                                name, cell, allowed_values_string, cell_label, val
                            )
                if label == headers[2]:
                    # for rules
                    allowed_values = (
                        config.abbr_no_beef
                        + config.abbr_no_fish
                        + config.abbr_no_chicken
                        + config.abbr_only_night
                        + config.abbr_only_day
                        + config.abbr_weekend_off
                        + config.abbr_weekend_on
                        + config.abbr_sunday_off
                        + config.abbr_sunday_on
                        + config.abbr_saturday_night_off
                    )
                    allowed_values_string = ", ".join(allowed_values)
                    sub_values = value.split(",")
                    sub_values = [val.strip() for val in sub_values]
                    for val in sub_values:
                        if val not in allowed_values:
                            cell = get_cell_for_boarder_from_df(
                                self.boarders_df, name, label
                            )
                            cell_label = label.replace("_", " - ")
                            self.add_to_invalid_cell(
                                name, cell, allowed_values_string, cell_label, val
                            )
                if "day" in label or "night" in label:
                    # for day and night | general meal
                    allowed_values = config.abbr_meal_on + config.abbr_meal_off
                    allowed_values_string = ", ".join(allowed_values)
                    sub_values = value.split(",")
                    sub_values = [val.strip() for val in sub_values]
                    for val in sub_values:
                        if val not in allowed_values:
                            cell = get_cell_for_boarder_from_df(
                                self.boarders_df, name, label
                            )
                            cell_label = label.replace("_", " - ")
                            self.add_to_invalid_cell(
                                name, cell, allowed_values_string, cell_label, val
                            )

    def get_total_meal_of_motnth(self):
        # get all the headers for meal counts
        headers = []
        for header in self.boarders_df_labels:
            if header.endswith("_day") or header.endswith("_night"):
                headers.append(header)
        meals_col_df = self.boarders_df[headers]
        all_val = meals_col_df.values
        total_meals = 0
        for val_row in all_val:
            for val in val_row:
                if not isinstance(val, str):
                    if math.isnan(val):
                        continue
                    return config.invalid_text
                if val not in config.abbr_meal_on + config.abbr_meal_off:
                    return config.invalid_text
                if val in config.abbr_meal_on:
                    total_meals += 1
        return total_meals

    def get_total_meal_of_day(self, date, time):
        # get header for meal count
        df_header = f"{date}_{time}"
        headers = [df_header]
        meals_col_df = self.boarders_df[headers]
        all_val = meals_col_df.values
        total_meals = 0
        for val_row in all_val:
            for val in val_row:
                if not isinstance(val, str):
                    if math.isnan(val):
                        continue
                    return config.invalid_text
                if val not in config.abbr_meal_on + config.abbr_meal_off:
                    return config.invalid_text
                if val in config.abbr_meal_on:
                    total_meals += 1
        guest_df_header = None
        for head in self.boarders_df.columns.tolist():
            if str(date) in head and time in head and "guest" in head:
                guest_df_header = head

        sum_guest_meal = 0
        if guest_df_header:
            sum_guest_meal = self.boarders_df[guest_df_header].sum()
        return total_meals + sum_guest_meal

    def get_total_guest_meals(self):
        headers = []
        for header in self.boarders_df_labels:
            if "guest" in header:
                if "grand" not in header:
                    headers.append(header)
        guest_meals_col_df = self.boarders_df[headers]
        contains_invalid = (
            guest_meals_col_df.map(lambda x: not isinstance(x, (float, int)))
            .any()
            .any()
        )
        if contains_invalid:
            return config.invalid_text
        total_guest = guest_meals_col_df.sum().sum()
        if total_guest == 0:
            return 0
        return total_guest

    def get_total_grand_guest_meals(self):
        headers = []
        for header in self.boarders_df_labels:
            if "guest" in header:
                if "grand" in header:
                    headers.append(header)
        if not len(headers):
            return 0
        guest_meals_col_df = self.boarders_df[headers]
        contains_invalid = (
            guest_meals_col_df.map(lambda x: not isinstance(x, (float, int)))
            .any()
            .any()
        )
        if contains_invalid:
            return config.invalid_text
        total_guest = guest_meals_col_df.sum().sum()
        if total_guest == 0:
            return 0
        return total_guest

    def get_total_deposits(self):
        # get all the headers for meal counts
        header = "Deposit"
        deposit_col_df = self.boarders_df[header]
        deposits = deposit_col_df.values
        for d in deposits:
            if not isinstance(d, (int, float)):
                return config.invalid_text

        total_deposit = deposit_col_df.sum()
        return total_deposit

    def see_low_budget_boarders(self):
        print("calculating data...")
        ok_boarders = []
        low_boarders = []
        no_deposit_boarders = []
        near_end_boarders = []
        for b in self.all_boarders:
            self.select_boarder(b, show_selected_msg=False)
            deposit = self.selected_boarder.deposit
            total_meal = self.selected_boarder.total_meal
            approx_charge = (
                total_meal * config.approx_meal_charge
            ) + config.approx_establishment_charge
            if deposit == 0:
                no_deposit_boarders.append([b, deposit, approx_charge])
                continue
            approx_charge -= config.threshold_amount_to_meal_off
            if deposit <= approx_charge:
                low_boarders.append([b, deposit, approx_charge])
            elif (
                deposit
                - (config.approx_meal_charge * config.thresold_meal_count_for_warning)
            ) <= approx_charge:
                near_end_boarders.append([b, deposit, approx_charge])
            elif deposit > approx_charge:
                ok_boarders.append([b, deposit, approx_charge])
        if len(no_deposit_boarders):
            print()
            print("BOARDERS WITH NO DEPOSIT:")
            for b, deposit, approx_charge in no_deposit_boarders:
                print(f"{b}, approx charge: ₹{round(approx_charge)}")
        if len(low_boarders):
            print()
            print("*Boarders with exceeded deposit amount:*")
            for b, deposit, approx_charge in low_boarders:
                print(
                    f"{b}, approx charge: ₹{round(approx_charge)} | deposit: ₹{deposit}"
                )
            print()
            print(
                f"_approx meal charge: {config.rupee_symbol}{config.approx_meal_charge}_"
            )
            print(
                f"_approx establishment charge: {config.rupee_symbol}{config.approx_establishment_charge}_"
            )
        if len(near_end_boarders):
            print()
            print("Boarders with near end amount:")
            for b, deposit, approx_charge in near_end_boarders:
                print(
                    f"{b}, approx charge: ₹{round(approx_charge)} | deposit: ₹{deposit}"
                )
        if len(ok_boarders):
            print()
            print("Boarders having good deposits:")
            for b, deposit, approx_charge in ok_boarders:
                print(
                    f"{b}, approx charge: ₹{round(approx_charge)} | deposit: ₹{deposit}"
                )
        print()
        input("press enter to continue...")

    def get_total_eggs(self):
        headers = []
        for header in self.boarders_df_labels:
            if "egg" in header:
                headers.append(header)
        eggs_col_df = self.boarders_df[headers]
        contains_invalid = (
            eggs_col_df.map(lambda x: not isinstance(x, (float, int))).any().any()
        )
        if contains_invalid:
            return config.invalid_text
        total_eggs = eggs_col_df.sum().sum()
        return total_eggs

    def print_meal_details(self):
        date_time = self.select_date_and_time(
            show_text="date for checking Meals", show_date_to_date=False
        )
        if date_time is None:
            return
        date_times = get_start_to_end_date_time(date_time)
        for date, time in date_times:
            n_dt, n_tm = get_date_time_next_meal_counting(self.month_days)
            if n_dt == date and n_tm == time:
                self.read_meal_details(date, time, for_msg=True)
            else:
                self.read_meal_details(date, time)

    def read_meal_details(self, date, time, for_msg=False) -> bool | None:
        print()
        current_date, _ = get_date_time_now()
        relavant_day = date
        if date == current_date:
            relavant_day = "today"
        elif date == current_date + 1:
            relavant_day = "tommorrow"
        text_meal = f"Boarders with Meal on({time}|{relavant_day}):"
        text_g_meal = f"Boarders with guest Meals({time}|{relavant_day}):"
        df_header = f"{date}_{time}"
        df_header_guest = ""
        for label in self.boarders_df_labels:
            if "guest" in label and time in label and f"{date}" in label:
                df_header_guest = label
                break
        meal_df = self.boarders_df[df_header]
        guest_meal_df = self.boarders_df[df_header_guest]
        is_blank = meal_df.map(lambda x: isinstance(x, float)).any()
        is_guest_blank = guest_meal_df.map(lambda x: not isinstance(x, int)).any()
        rules_df = self.boarders_df["Rules"]

        if is_blank or is_guest_blank:
            print(
                f"{config.warning_string}Some cells are empty on: {time}|{relavant_day}"
            )
            print()
        meals = list(
            zip(
                self.all_boarders, meal_df.values, guest_meal_df.values, rules_df.values
            )
        )

        menu_item = None
        if for_msg:
            menu_item = select_a_item_with_default(
                self.all_routine_values,
                default=self.get_next_meal_routine().lower(),
                presentation_text="Select Meal-menu type:",
                presentation_symbol="_",
                input_text="Meal Item",
                quit_text="DONT CONSIDER!",
            )

        on_meals = []
        on_guest_meals = []
        total_meals = 0
        meals_to_alternate = 0
        for meal in meals:
            name, n_meal, g_meal, rule = meal
            if n_meal in config.abbr_meal_on:
                f_name = name
                if menu_item and isinstance(rule, str):
                    if check_meal_alternate(menu_item, rule):
                        f_name += f" (_{menu_item} off_)"
                        meals_to_alternate += 1
                on_meals.append(f_name)
                total_meals += 1
            if isinstance(g_meal, int):
                if g_meal > 0:
                    on_guest_meals.append(f"{name}: {g_meal}")
                    total_meals += g_meal
        if not len(on_meals + on_guest_meals):
            print()
            print(f"{config.info_string}NO MEAL IS ON: {time}|{relavant_day}")
            print()
            input("press Enter to continue...")
            return
        on_meals = sorted(on_meals)
        on_guest_meals = sorted(on_guest_meals)
        if for_msg:
            print()
            print("Boarders having meal-on:")
            print()
            print(f"```date: {date}-{self.month_string}-{self.year}|{time}```")
            for i, meal in enumerate(on_meals, start=1):
                print(f"{i}. {meal}")
            print()
            print("Guest Meals:")
            if len(on_guest_meals):
                for meal in on_guest_meals:
                    print(f"  {meal}")
            else:
                print("   No Guest Meal is On.")
            print()
            print(f"*Total meal: {total_meals}*")
            print()
            if meals_to_alternate:
                print(f"*_No {menu_item}: {meals_to_alternate}_*")
            print()
            print(
                "⚠️  _Karo Meal On/Off hole obossoi niche Total meal mention kore debe_"
            )
            print()
            input("press Enter to continue...")
        else:
            on_meals = [f"   {meal}" for meal in on_meals]
            on_guest_meals = [f"   {meal}" for meal in on_guest_meals]
            print(text_meal)
            print_in_2_cols(on_meals)
            print()
            print(text_g_meal)
            print_in_2_cols(on_guest_meals)
            input("press Enter to continue...")

    def get_total_accessories_charge(self):
        headers = ["Acs Charge"]
        acs_charge_col_df = self.boarders_df[headers]
        contains_invalid = (
            acs_charge_col_df.map(lambda x: not isinstance(x, (float, int))).any().any()
        )
        if contains_invalid:
            return config.invalid_text
        total_acs_charge = acs_charge_col_df.sum().sum()
        return total_acs_charge

    def calculate_meal_charge(self):
        # ( total_expenditure - (guest_meal_charge + eggs_charge )) / totoal boarders
        total_expenditure = (
            self.total_marketing + self.total_extra_rice + self.total_extra_marketing
        )
        minus_amount = (
            (self.total_guest_meals * config.guest_meal_price)
            + (self.total_grand_guest_meals * config.grand_guest_price)
            + (self.total_eggs * config.egg_price)
        )
        self.meal_charge = (total_expenditure - minus_amount) / self.total_meals
        return self.meal_charge

    def calculate_establishment_charge(self):
        excel_file = self.filename
        self.establishment_df = pd.read_excel(
            io=excel_file,
            sheet_name=config.establishment_sheet_name,
            engine="openpyxl",
            header=1,
            usecols=range(2),
        )
        # estd = (total estd amount - accesoris charge) / total boarders
        self.total_establishment_amount = self.establishment_df["Amount"].sum()
        self.establishment_charge = (
            self.total_establishment_amount - self.get_total_accessories_charge()
        ) / self.total_boarders

        return self.establishment_charge

    def get_informations_to_show(self) -> list:
        informations = []
        current_date, _ = get_date_time_now()
        date, time = get_date_time_next_meal_counting(self.month_days)
        relavant_day = date
        if date == current_date:
            relavant_day = "today"
        elif date == current_date + 1:
            relavant_day = "tommorrow"
        # get informations
        meal_counting = self.allow_next_meal_count()
        if meal_counting is not None:
            if meal_counting is False:
                informations.append(
                    f"meal count of {relavant_day}({time}): {self.do_next_meal_count()}"
                )
            elif meal_counting:
                informations.append(meal_counting)

        informations.append(f"Total Boarders: {self.total_boarders}")
        informations.append(f"Total Meals: {self.get_total_meal_of_motnth()}")
        informations.append(
            f"Total Deposits: {config.rupee_symbol}{self.get_total_deposits()}"
        )
        total_guest_meals = self.get_total_guest_meals()
        if total_guest_meals:
            informations.append(f"Total Guest Meals: {total_guest_meals}")
        total_grand_guest_meals = self.get_total_grand_guest_meals()
        if total_grand_guest_meals:
            informations.append(f"Total grand Guest Meals: {total_grand_guest_meals}")
        if self.total_eggs:
            informations.append(f"Total extra eggs: {self.total_eggs}")
        if self.total_extra_rice:
            informations.append(
                f"Total extra rice: {config.rupee_symbol}{self.total_extra_rice}"
            )
        if self.total_marketing:
            informations.append(
                f"Total Marketing: {config.rupee_symbol}{self.total_marketing}"
            )
        if self.total_extra_marketing:
            informations.append(
                f"Total extra Marketing: {config.rupee_symbol}{self.total_extra_marketing}"
            )
        if self.total_marketing or self.total_extra_marketing or self.total_extra_rice:
            informations.append(
                f"Remainig amount: {config.rupee_symbol}{self.get_total_deposits() - (self.total_marketing + self.total_extra_marketing +  self.total_extra_rice)}"
            )
        if self.total_establishment_amount:
            informations.append(
                f"Total estd amount: {config.rupee_symbol}{self.total_establishment_amount}"
            )
            informations.append(
                f"Estd Charge: {config.rupee_symbol}{self.establishment_charge:.2f}"
            )
        if self.meal_charge > 0:
            informations.append(
                f"Current meal charge: {config.rupee_symbol}{self.meal_charge:.2f}"
            )

        # return informations list
        return informations

    ############# Editing
    def set_grand_meal_day(self):
        columns = list(self.boarders_df.columns)
        default_date = int(columns[-1].split("_")[0])
        chosen_date = choose_a_date(
            self.month_days, self.month_string, default=default_date
        )
        if chosen_date is None:
            return
        chosen_time = choose_time()
        if chosen_time is None:
            return
        in_night = True if chosen_time == "night" else False
        in_day = True if chosen_time == "day" else False

        def filter_day_label(label):
            if label.startswith(f"{chosen_date}_") and "guest" in label:
                if in_night:
                    if "night" in label:
                        return True
                if in_day:
                    if "day" in label:
                        return True
            return False

        grand_labels = list(filter(filter_day_label, columns))
        if len(grand_labels) != 1:
            print(
                f"{config.failure_string}multiple dates found!! cannot set grandmeal day."
            )
            return
        grand_label = grand_labels[-1]
        grand_col_label = get_cell_for_boarder_from_df(
            self.boarders_df, None, grand_label, only_header=True
        )
        value = get_val_from_excel_file(config.meal_count_sheet_name, grand_col_label)
        if "grand_" not in value:
            value = "grand_" + value
        else:
            print(f"{config.success_string}already marked as grand meal {chosen_time}.")
        write_in_excel_file(
            config.meal_count_sheet_name,
            grand_col_label,
            value,
            fill=grand_meal_fill,
            color=grand_meal_color,
        )
        # give confirmation message
        print(
            f"{config.success_string}{chosen_date}-{chosen_time} successfully marked as a grand meal {chosen_time}."
        )

    def remove_grand_meal_day(self):
        in_night = True
        in_day = True

        def filter_day_label(label):
            if "grand_" in label and "guest" in label:
                if in_night:
                    if "night" in label:
                        return True
                if in_day:
                    if "day" in label:
                        return True
            return False

        columns = list(self.boarders_df.columns)
        grand_labels = list(filter(filter_day_label, columns))
        if len(grand_labels) == 0:
            print(f"{config.info_string} No dates marked as a grandmeal day.")
            input()
            return

        grand_labels = select_items_from_list(
            grand_labels,
            presentation_text="Remove Grand-Meal mark from:",
            confirm_on_all_selection=True,
        )
        if grand_labels is None:
            return
        for grand_label in grand_labels:
            grand_col_label = get_cell_for_boarder_from_df(
                self.boarders_df, None, grand_label, only_header=True
            )
            value = get_val_from_excel_file(
                config.meal_count_sheet_name, grand_col_label
            )
            if "grand_" in value:
                value = value.replace("grand_", "")
            chosen_date = int(grand_label.split("_")[0])
            chosen_time = grand_label.split("(")[-1].split(")")[0]
            write_in_excel_file(
                config.meal_count_sheet_name,
                grand_col_label,
                value,
                fill=get_date_fill(chosen_date),
                color=get_date_color(chosen_date),
            )
            # give confirmation
            print(
                f"{config.success_string}{chosen_date}-{chosen_time} successfully unmarked from grand-meal."
            )

    def add_boarder(self, name=None):
        if name is None:
            name = take_input("Enter Boarder name:", confirm=True)
            if name is None:
                print("cancelled: Add boarder.")
                return
            name = name.strip()
        while self.select_boarder(name, select=False):
            name = input(
                f"{config.info_string}A boarder with name '{name}' already exist!! try with other name: "
            )
        excel_file = self.filename
        work_book = load_workbook(excel_file)
        work_sheet = work_book[config.meal_count_sheet_name]
        insertion_row = (self.boarders_df.index[-1]) + 1
        write_to_cell_excel(
            work_sheet, f"A{insertion_row}", name.title(), center_aligned=True
        )
        work_book.save(self.filename)
        print(f"{config.success_string}'{name.title()}' added to Boarder list.")
        input()

    ########### Helpers
    def show_invalid_cells(self):
        print()
        if not len(self.invalid_cells[:-1]):
            print(
                f"{config.success_string} All values are valid in {config.workbook_name} file."
            )
            input()
            return
        invalid_cells = self.invalid_cells[:-1]
        # invalid_cells = sorted(invalid_cells)
        print(f"{config.failure_string} Invalid cells are:")
        for inv in invalid_cells:
            print(f"-> {inv['name']}")
            for inv_txt in inv["cells"]:
                print(f"    {inv_txt}")
        input()

    def allow_next_meal_count(
        self,
        respect_month=config.respect_month_for_meal_counting,
        respect_year=config.respect_year_for_meal_counting,
    ) -> bool | None:
        calculate = True
        if respect_year:
            if self.year != pendulum.now().year:
                calculate = False
        if respect_month:
            if pendulum.now().month != self.month:
                calculate = False

        if not calculate:
            return None

        date, time = get_date_time_next_meal_counting(self.month_days)
        current_date, _ = get_date_time_now()
        relavant_day = date
        if date == current_date:
            relavant_day = "today"
        elif date == current_date + 1:
            relavant_day = "tommorrow"
        text = f"Pending Meal Count for {time}({relavant_day})"
        df_header = f"{date}_{time}"
        meal_df = self.boarders_df[df_header]
        is_blank = meal_df.map(lambda x: isinstance(x, float)).any()
        if is_blank:
            return text
        return False

    def get_next_meal_routine(self) -> str:
        date, time = get_date_time_next_meal_counting(self.month_days)
        day = pendulum.date(self.year, self.month, date).format("dddd").lower()
        return self.routine[self.routine[config.routine[0][0]] == day][time].values[0]

    def do_next_meal_count(self):
        allowed = self.allow_next_meal_count()
        # allowed = True
        if allowed is None:
            return
        # date, time = 2, "night"
        date, time = get_date_time_next_meal_counting(self.month_days)
        if allowed is False:
            # get total meal
            return self.get_total_meal_of_day(date, time)
        for boarder in self.all_boarders:
            self.select_boarder(boarder, show_selected_msg=False)
            self.selected_boarder.turn_on_meal(date, time, follow_rule=True)

    def add_to_invalid_cell(
        self, name, invalid_cell, valid_format, cell_label=None, value=None
    ):
        if cell_label is not None:
            cell_label = f" [{cell_label}] "
        else:
            cell_label = ""

        if value is not None:
            value = f" value: {value} |"
        else:
            value = ""

        final_invalid_cells = []
        invalid_cells = [[]]
        if len(self.invalid_cells):
            invalid_cells = self.invalid_cells

        # {name: , cells: ['cell - correct form']}
        cell = {}
        found = False
        for inv_cell in invalid_cells[:-1]:
            if inv_cell["name"] == name:
                found = True
                cell["name"] = name
                cell["cells"] = inv_cell["cells"]
                cell["cells"].append(
                    f"{invalid_cell}{cell_label} |{value} Allowed: {valid_format}"
                )
                final_invalid_cells.append(cell)
            else:
                final_invalid_cells.append(inv_cell)

        if not found:
            final_invalid_cells.append(
                {
                    "name": name,
                    "cells": [
                        f"{invalid_cell}{cell_label} |{value} Allowed: {valid_format}"
                    ],
                }
            )

        # last item is all the invalid cells
        cells = invalid_cells[-1]
        if invalid_cell not in cells:
            cells.append(invalid_cell)

        final_invalid_cells.append(cells)
        self.invalid_cells = final_invalid_cells

    def select_boarder(self, name=None, select=True, show_selected_msg=True):
        # get Names in a list
        columns = list(self.boarders_df.columns)
        names_list = self.boarders_df[columns[0]].tolist()

        choice = None
        select_a_boarder = True
        if name is not None:
            if name.lower() in [n.strip().lower() for n in names_list]:
                choice = name
                select_a_boarder = False
            else:
                return False

        if select_a_boarder:
            presentation_text = "Select a Boarder:"
            choice = fuzzy_select_from_list(
                names_list,
                sort=True,
                presentation_symbol="-",
                presentation_text=presentation_text,
            )
        if choice is None:
            return
        if select:
            selected_row = self.boarders_df[self.boarders_df[columns[0]] == choice]
            self.selected_boarder_row = (selected_row.index)[0]
            # marketing row
            if self.marketing_df.empty:
                selected_marketing = None
            else:
                selected_marketing = self.marketing_df[
                    self.marketing_df["Name"] == choice
                ]
                if selected_marketing.empty:
                    selected_marketing = None
            # extra marketing row
            if self.extra_marketing_df.empty:
                selected_extra_marketing = None
            else:
                selected_extra_marketing = self.extra_marketing_df[
                    self.extra_marketing_df["Name"] == choice
                ]
                if selected_extra_marketing.empty:
                    selected_extra_marketing = None

            if show_selected_msg:
                print()
                print(f"{config.selected_string}'{choice}' IS NOW SELECTED.")
            self.selected_boarder = Boarder(
                selected_row,
                invalid_cells=self.invalid_cells[-1],
                month=self.month,
                year=self.year,
                marketing_df=selected_marketing,
                extra_markeing_df=selected_extra_marketing,
                file_to_work_on=self.filename,
            )
        return True

    def select_multiple_boarders(self, fuzzy: bool = False):
        # get Names in a list
        columns = list(self.boarders_df.columns)
        names_list = self.boarders_df[columns[0]].tolist()
        selected_boarders = None
        if not fuzzy:
            presentation_text = "Select Boarders:"
            selected_boarders = select_items_from_list(
                names_list,
                sort=True,
                presentation_symbol="-",
                presentation_text=presentation_text,
                allow_all_selection=config.allow_all_boarder_selection,
                confirm_on_all_selection=True,
            )
        else:
            presentation_text = "Select Boarders: (fuzzy)"
            selected_boarders = fuzzy_select_from_list(
                names_list,
                multi_select=True,
                sort=True,
                presentation_symbol="-",
                presentation_text=presentation_text,
            )
        if selected_boarders is None:
            return
        else:
            return selected_boarders

    def select_date_and_time(
        self,
        month=None,
        year=None,
        show_tomorrow=True,
        show_yesterday=True,
        show_date_to_date=True,
        show_next_meal_count=True,
        show_text="date",
    ):
        if month is None:
            month = self.month
        if year is None:
            year = self.year
        date = None
        date_today = int(pendulum.today().format("DD"))
        day_today = pendulum.today().format("dddd")
        select_next_meal_str = f"last meal count"
        select_date_str = "choose a date"
        select_date_to_date_str = "Date to Date"
        today_text = f"today ({day_today})"
        tomorrow_text = "tomorrow"
        yesterday_text = "yesterday"
        curr_month = pendulum.now().month
        curr_year = pendulum.now().year
        options = []
        if month == curr_month and year == curr_year:
            if show_next_meal_count and self.allow_next_meal_count() is False:
                options.append(select_next_meal_str)
            options.append(today_text)
            if show_tomorrow and date_today < self.month_days - 1:
                options.append(tomorrow_text)
            if show_yesterday and not date_today < 2:
                options.append(yesterday_text)
        options.append(select_date_str)
        if show_date_to_date:
            options.append(select_date_to_date_str)
        choice = select_items_from_list(
            options,
            multi_select=False,
            quit_text="CANCEL",
            presentation_text=f"Choose a {show_text} | today: {date_today}, {self.month_string.title()}",
        )
        if choice is None:
            return
        if choice == select_next_meal_str:
            return get_date_time_next_meal_counting(self.month_days)
        if choice == today_text:
            date = date_today
        if choice == tomorrow_text:
            date = int(pendulum.tomorrow().format("DD"))
        if choice == yesterday_text:
            date = int(pendulum.yesterday().format("DD"))
        if choice == select_date_str:
            date = choose_a_date(
                self.month_days,
                self.month_string,
                default=int(pendulum.today().format("DD")),
            )
        if choice == select_date_to_date_str:
            return self.select_date_to_date()
        if date is None:
            return

        time = None
        time = choose_time()
        if time is None:
            return

        return (date, time)

    def select_date_to_date(self):
        date_today = int(pendulum.today().format("DD"))
        day_today = pendulum.today().format("dddd")
        from_date = choose_a_date(
            self.month_days,
            self.month_string,
            default=date_today,
            show_text=f"START Date | today: {date_today}, {day_today}",
        )
        if from_date is None:
            return
        choose_time_str = "start time:"
        from_time = choose_time(show_text=choose_time_str)
        if from_time is None:
            return

        to_date = choose_a_date(
            self.month_days,
            self.month_string,
            default=self.month_days,
            show_text=f"END Date | today: {date_today}, {day_today}",
        )
        if to_date is None:
            return
        choose_time_str = "end time:"
        to_time = choose_time(show_text=choose_time_str)
        if to_time is None:
            return
        return ((from_date, from_time), (to_date, to_time))

    ## writing final sheet
    def write_final_sheet(self):
        # preparing file
        sheet = "final chart"
        workbook = load_workbook(self.filename)
        if sheet not in workbook.sheetnames:
            worksheet = workbook.create_sheet(title=sheet)
        else:
            worksheet = workbook[sheet]
            workbook.remove(worksheet)
            worksheet = workbook.create_sheet(title=sheet)
        # write headers
        headers = [
            "SL NO.",
            "NAME",
            "TOTAL MEAL",
            "EGGS",
            "GUEST MEAL",
            "GUEST MEAL CHARGE",
            "ACS.(comp, fan etc) CHARGE",
            "TOTAL CHARGE",
            "DEPOSIT",
            "DUE",
            "REFUND",
            "PAYMENT DATE",
        ]
        guest_meals_text = ["REGULAR", "GRAND"]
        guest_meals_text.reverse()
        guest_idx = headers.index("GUEST MEAL")
        headers.pop(guest_idx)
        for txt in guest_meals_text:
            headers.insert(guest_idx, txt)
        guest_meals_idx = [guest_idx + i + 1 for i in range(len(guest_meals_text))]
        end_label = get_excel_column_label(len(headers))
        worksheet.merge_cells(f"A1:{end_label}1")
        write_to_cell_excel(worksheet, "A1", "Meal Charge Sheet", center_aligned=True)
        worksheet.merge_cells(f"A2:{end_label}2")
        write_to_cell_excel(
            worksheet,
            "A2",
            f"Month: {self.month_string.title()}",
            bold=True,
            center_aligned=True,
        )
        worksheet.merge_cells(f"A3:{end_label}3")
        write_to_cell_excel(
            worksheet,
            "A3",
            f"Manager: {config.MANAGER_NAME}",
            bold=True,
            center_aligned=True,
        )

        done_guest_header = False
        for i, val in enumerate(headers, start=1):
            label = get_excel_column_label(i)
            cell = f"{label}5"
            bold = True
            if i in guest_meals_idx and not done_guest_header:
                worksheet.merge_cells(
                    f"{get_excel_column_label(guest_meals_idx[0])}5:{get_excel_column_label(guest_meals_idx[-1])}5"
                )
                write_to_cell_excel(
                    worksheet, cell, "GUEST MEAL", bold=True, center_aligned=True
                )

                done_guest_header = True
            if i in guest_meals_idx:
                cell = f"{label}6"
                bold = False
            else:
                worksheet.merge_cells(f"{cell}:{label}6")
            write_to_cell_excel(worksheet, cell, val, bold=bold, center_aligned=True)

        # Get all columns
        sl_no_col = get_excel_column_label(headers.index("SL NO.") + 1)
        name_col = get_excel_column_label(headers.index("NAME") + 1)
        meals_col = get_excel_column_label(headers.index("TOTAL MEAL") + 1)
        eggs_col = get_excel_column_label(headers.index("EGGS") + 1)
        regular_guest_col = get_excel_column_label(headers.index("REGULAR") + 1)
        grand_guest_col = get_excel_column_label(headers.index("GRAND") + 1)
        total_guest_charge_col = get_excel_column_label(
            headers.index("GUEST MEAL CHARGE") + 1
        )
        acs_charge_col = get_excel_column_label(
            headers.index("ACS.(comp, fan etc) CHARGE") + 1
        )
        total_charge_col = get_excel_column_label(headers.index("TOTAL CHARGE") + 1)
        deposit_col = get_excel_column_label(headers.index("DEPOSIT") + 1)
        due_col = get_excel_column_label(headers.index("DUE") + 1)
        refund_col = get_excel_column_label(headers.index("REFUND") + 1)

        # write all boarders
        last_row = 1
        for i, b in enumerate(self.all_boarders, start=1):
            row = 6 + i
            last_row = row
            due = None
            refund = None
            self.select_boarder(b, show_selected_msg=False)
            guest_meal_charge = (
                self.selected_boarder.guest_meals * config.guest_meal_price
            ) + (self.selected_boarder.grand_guest_meals * config.grand_guest_price)
            total_charge = (
                (self.selected_boarder.total_meal * self.meal_charge)
                + (self.selected_boarder.guest_meals * config.guest_meal_price)
                + (self.selected_boarder.total_eggs * config.egg_price)
                + (self.selected_boarder.grand_guest_meals * config.grand_guest_price)
                + self.selected_boarder.accessories_charge
                + self.establishment_charge
            )
            write_to_cell_excel(
                worksheet,
                f"{sl_no_col}{row}",
                i,
                center_aligned=True,
            )
            write_to_cell_excel(
                worksheet,
                f"{name_col}{row}",
                self.selected_boarder.name,
                center_aligned=True,
            )
            write_to_cell_excel(
                worksheet,
                f"{meals_col}{row}",
                self.selected_boarder.total_meal,
                center_aligned=True,
            )
            if self.selected_boarder.total_eggs:
                write_to_cell_excel(
                    worksheet,
                    f"{eggs_col}{row}",
                    self.selected_boarder.total_eggs,
                    center_aligned=True,
                )
            if self.selected_boarder.guest_meals:
                write_to_cell_excel(
                    worksheet,
                    f"{regular_guest_col}{row}",
                    self.selected_boarder.guest_meals,
                    center_aligned=True,
                )
            if self.selected_boarder.grand_guest_meals:
                write_to_cell_excel(
                    worksheet,
                    f"{grand_guest_col}{row}",
                    self.selected_boarder.grand_guest_meals,
                    center_aligned=True,
                )
            if guest_meal_charge:
                write_to_cell_excel(
                    worksheet,
                    f"{total_guest_charge_col}{row}",
                    f"{config.rupee_symbol}{guest_meal_charge}",
                    center_aligned=True,
                )
            if self.selected_boarder.accessories_charge:
                write_to_cell_excel(
                    worksheet,
                    f"{acs_charge_col}{row}",
                    f"{config.rupee_symbol}{self.selected_boarder.accessories_charge}",
                    center_aligned=True,
                )
            if total_charge:
                write_to_cell_excel(
                    worksheet,
                    f"{total_charge_col}{row}",
                    f"{config.rupee_symbol}{total_charge:.2f}",
                    center_aligned=True,
                )
            write_to_cell_excel(
                worksheet,
                f"{deposit_col}{row}",
                f"{config.rupee_symbol}{self.selected_boarder.deposit}",
                center_aligned=True,
            )
            remaining = total_charge - self.selected_boarder.deposit
            if remaining > 0:
                due = remaining
            if remaining < 0:
                refund = abs(remaining)
            if due is not None:
                write_to_cell_excel(
                    worksheet,
                    f"{due_col}{row}",
                    f"{config.rupee_symbol}{due:.2f}",
                    bold=True,
                    center_aligned=True,
                )

            if refund is not None:
                write_to_cell_excel(
                    worksheet,
                    f"{refund_col}{row}",
                    f"{config.rupee_symbol}{refund:.2f}",
                    center_aligned=True,
                )
        # write meal charge and establishment charge
        last_row += 1
        meal_charge_row = last_row + 1
        if self.meal_charge:
            worksheet.merge_cells(f"A{meal_charge_row}:{end_label}{meal_charge_row}")
            write_to_cell_excel(
                worksheet,
                f"A{meal_charge_row}",
                f"Meal Charge: {config.rupee_symbol}{self.meal_charge:.2f}",
                bold=True,
                center_aligned=True,
            )
        estd_charge_row = last_row + 2
        if self.establishment_charge:
            worksheet.merge_cells(f"A{estd_charge_row}:{end_label}{estd_charge_row}")
            write_to_cell_excel(
                worksheet,
                f"A{estd_charge_row}",
                f"Establishment Charge: {config.rupee_symbol}{self.establishment_charge:.2f}",
                bold=True,
                center_aligned=True,
            )

        workbook.save(self.filename)
        workbook.close()
        print(f"{config.success_string}Done producing final chart, sheet name: {sheet}")
        print()
        input("press Enter to continue...")
