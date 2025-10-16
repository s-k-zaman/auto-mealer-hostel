from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
import config


def get_excel_column_label(column_number):
    column_label = ""
    while column_number > 0:
        column_number, remainder = divmod(column_number - 1, 26)
        column_label = chr(65 + remainder) + column_label
    return column_label


def is_sheet_present_excel(work_sheet, work_file=config.workbook_name):
    workbook = load_workbook(work_file)
    worksheet_name = work_sheet
    # Check if the worksheet already exists
    sheet_names = workbook.sheetnames
    workbook.close()
    if worksheet_name in sheet_names:
        return True
    return False


def write_to_cell_excel(
    sheet,
    cell,
    value,
    bold=False,
    center_aligned=False,
    right_aligned=False,
    color=False,
    color_reset=False,
    fill=False,
    fill_reset=False,
):
    sheet[cell] = value
    if bold:
        sheet[cell].font = Font(bold=True)
    if center_aligned:
        sheet[cell].alignment = Alignment(horizontal="center", vertical="center")
    if right_aligned:
        sheet[cell].alignment = Alignment(horizontal="right", vertical="center")
    if fill:
        fill = fill.replace("#", "")
        fill_with = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
        sheet[cell].fill = fill_with
    if fill_reset:
        fill_with = PatternFill(fill_type="none")
        sheet[cell].fill = fill_with
    if color:
        color = color.replace("#", "")
        sheet[cell].font = Font(color=color)
    if color_reset:
        sheet[cell].font = Font(color=None)


def get_val_from_excel_file(sheet, cell, filename=config.workbook_name):
    ## preparing file
    work_file = load_workbook(filename)
    if sheet not in work_file.sheetnames:
        # If it doesn't exist
        print(f"no sheet named `{sheet}` found.")
        return False
    work_sheet = work_file[sheet]
    ## reading vaule
    value = work_sheet[cell].value
    work_file.close()
    return value


def write_in_excel_file(sheet, cell, val, filename=config.workbook_name, **kwargs):
    ## preparing file
    work_file = load_workbook(filename)
    if sheet not in work_file.sheetnames:
        # If it doesn't exist
        print(f"no sheet named `{sheet}` found.")
        return False
    work_sheet = work_file[sheet]
    ## editing the sheet
    write_to_cell_excel(work_sheet, cell, val, **kwargs)
    work_file.save(filename)
    work_file.close()
    return True


def get_cell_for_boarder_from_df(
    boarders_df, boarder, header, only_header=False
) -> str:
    row = None
    if only_header:
        row = 2
    else:
        selected_row = boarders_df[boarders_df["Name"] == boarder]
        if len(selected_row) < 1:
            raise Exception(f'"{boarder}" is not present')
        if len(selected_row) > 1:
            raise Exception(f"more than one boarder present with name: {boarder}")
        row = (selected_row.index)[0]

    headers = list(boarders_df.columns)
    if header not in headers:
        raise Exception(f"{header} is not a valid header")
    idx = headers.index(header)
    col_num = idx + 1
    column = get_excel_column_label(col_num)
    return f"{column}{row}"
